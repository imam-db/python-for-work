"""
Excel Comparator - Membandingkan 2 file Excel dan menampilkan perbedaannya
Mendukung konfigurasi via YAML dan CLI arguments (hybrid)
"""

import sys
import argparse
import pandas as pd
import yaml
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


# Warna untuk highlight
FILL_ADDED = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Hijau
FILL_DELETED = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Merah
FILL_CHANGED = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Kuning


def load_config(config_path: str) -> dict:
    """Load konfigurasi dari file YAML."""
    config_file = Path(config_path)
    if config_file.exists():
        with open(config_file, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f) or {}
    return {}


def compare_dataframes(df_old: pd.DataFrame, df_new: pd.DataFrame, key_column: str = None) -> dict:
    """
    Bandingkan dua DataFrame dan return perbedaannya.
    
    Returns:
        dict dengan keys: added, deleted, changed
    """
    result = {
        'added': [],      # Baris baru
        'deleted': [],    # Baris dihapus
        'changed': [],    # Cell yang berubah
        'summary': {}
    }
    
    if key_column and key_column in df_old.columns and key_column in df_new.columns:
        # Bandingkan berdasarkan key column
        old_keys = set(df_old[key_column].astype(str))
        new_keys = set(df_new[key_column].astype(str))
        
        # Baris baru (ada di new, tidak di old)
        added_keys = new_keys - old_keys
        for key in added_keys:
            row = df_new[df_new[key_column].astype(str) == key].iloc[0].to_dict()
            result['added'].append({'key': key, 'data': row})
        
        # Baris dihapus (ada di old, tidak di new)
        deleted_keys = old_keys - new_keys
        for key in deleted_keys:
            row = df_old[df_old[key_column].astype(str) == key].iloc[0].to_dict()
            result['deleted'].append({'key': key, 'data': row})
        
        # Baris yang sama, cek perubahan nilai
        common_keys = old_keys & new_keys
        for key in common_keys:
            old_row = df_old[df_old[key_column].astype(str) == key].iloc[0]
            new_row = df_new[df_new[key_column].astype(str) == key].iloc[0]
            
            for col in df_old.columns:
                if col in df_new.columns:
                    old_val = old_row[col]
                    new_val = new_row[col]
                    
                    # Handle NaN comparison
                    if pd.isna(old_val) and pd.isna(new_val):
                        continue
                    if str(old_val) != str(new_val):
                        result['changed'].append({
                            'key': key,
                            'column': col,
                            'old_value': old_val,
                            'new_value': new_val
                        })
    else:
        # Bandingkan berdasarkan posisi baris
        max_rows = max(len(df_old), len(df_new))
        
        for i in range(max_rows):
            if i >= len(df_old):
                # Baris baru
                result['added'].append({'row': i + 1, 'data': df_new.iloc[i].to_dict()})
            elif i >= len(df_new):
                # Baris dihapus
                result['deleted'].append({'row': i + 1, 'data': df_old.iloc[i].to_dict()})
            else:
                # Bandingkan cell by cell
                for col in df_old.columns:
                    if col in df_new.columns:
                        old_val = df_old.iloc[i][col]
                        new_val = df_new.iloc[i][col]
                        
                        if pd.isna(old_val) and pd.isna(new_val):
                            continue
                        if str(old_val) != str(new_val):
                            result['changed'].append({
                                'row': i + 1,
                                'column': col,
                                'old_value': old_val,
                                'new_value': new_val
                            })
    
    result['summary'] = {
        'added_count': len(result['added']),
        'deleted_count': len(result['deleted']),
        'changed_count': len(result['changed'])
    }
    
    return result


def print_report(result: dict, key_column: str = None):
    """Print laporan perbandingan ke console."""
    summary = result['summary']
    
    print("\n" + "=" * 50)
    print("LAPORAN PERBANDINGAN")
    print("=" * 50)
    
    # Summary
    print(f"\nRingkasan:")
    print(f"  - Baris baru     : {summary['added_count']}")
    print(f"  - Baris dihapus  : {summary['deleted_count']}")
    print(f"  - Cell berubah   : {summary['changed_count']}")
    
    if summary['added_count'] == 0 and summary['deleted_count'] == 0 and summary['changed_count'] == 0:
        print("\n✓ Kedua file IDENTIK, tidak ada perbedaan.")
        return
    
    # Detail perubahan
    if result['changed']:
        print(f"\n--- PERUBAHAN ({len(result['changed'])}) ---")
        for item in result['changed'][:20]:  # Limit 20
            if key_column:
                print(f"  [{key_column}={item['key']}] {item['column']}: {item['old_value']} → {item['new_value']}")
            else:
                print(f"  Baris {item['row']}, {item['column']}: {item['old_value']} → {item['new_value']}")
        if len(result['changed']) > 20:
            print(f"  ... dan {len(result['changed']) - 20} perubahan lainnya")
    
    if result['added']:
        print(f"\n--- BARIS BARU ({len(result['added'])}) ---")
        for item in result['added'][:10]:
            if key_column:
                print(f"  [{key_column}={item['key']}]")
            else:
                print(f"  Baris {item['row']}")
        if len(result['added']) > 10:
            print(f"  ... dan {len(result['added']) - 10} baris lainnya")
    
    if result['deleted']:
        print(f"\n--- BARIS DIHAPUS ({len(result['deleted'])}) ---")
        for item in result['deleted'][:10]:
            if key_column:
                print(f"  [{key_column}={item['key']}]")
            else:
                print(f"  Baris {item['row']}")
        if len(result['deleted']) > 10:
            print(f"  ... dan {len(result['deleted']) - 10} baris lainnya")


def export_to_excel(df_old: pd.DataFrame, df_new: pd.DataFrame, result: dict, 
                    output_path: str, key_column: str = None):
    """Export hasil perbandingan ke Excel dengan highlight."""
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary['A1'] = "Laporan Perbandingan Excel"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A3'] = "Baris Baru:"
    ws_summary['B3'] = result['summary']['added_count']
    ws_summary['A4'] = "Baris Dihapus:"
    ws_summary['B4'] = result['summary']['deleted_count']
    ws_summary['A5'] = "Cell Berubah:"
    ws_summary['B5'] = result['summary']['changed_count']
    
    # Sheet 2: Detail Perubahan
    if result['changed']:
        ws_changed = wb.create_sheet("Perubahan")
        if key_column:
            ws_changed.append([key_column, "Kolom", "Nilai Lama", "Nilai Baru"])
        else:
            ws_changed.append(["Baris", "Kolom", "Nilai Lama", "Nilai Baru"])
        
        for item in result['changed']:
            if key_column:
                ws_changed.append([item['key'], item['column'], str(item['old_value']), str(item['new_value'])])
            else:
                ws_changed.append([item['row'], item['column'], str(item['old_value']), str(item['new_value'])])
        
        # Highlight header
        for cell in ws_changed[1]:
            cell.fill = FILL_CHANGED
            cell.font = Font(bold=True)
    
    # Sheet 3: Baris Baru
    if result['added']:
        ws_added = wb.create_sheet("Baris Baru")
        # Header dari data
        if result['added']:
            headers = list(result['added'][0]['data'].keys())
            ws_added.append(headers)
            for cell in ws_added[1]:
                cell.fill = FILL_ADDED
                cell.font = Font(bold=True)
            
            for item in result['added']:
                ws_added.append(list(item['data'].values()))
    
    # Sheet 4: Baris Dihapus
    if result['deleted']:
        ws_deleted = wb.create_sheet("Baris Dihapus")
        if result['deleted']:
            headers = list(result['deleted'][0]['data'].keys())
            ws_deleted.append(headers)
            for cell in ws_deleted[1]:
                cell.fill = FILL_DELETED
                cell.font = Font(bold=True)
            
            for item in result['deleted']:
                ws_deleted.append(list(item['data'].values()))
    
    wb.save(output_path)
    print(f"\nLaporan Excel disimpan ke: {output_path}")


def compare_excel(file_old: str, file_new: str, key_column: str = None, 
                  output_file: str = None) -> dict:
    """Main function untuk membandingkan 2 file Excel."""
    
    # Validasi file
    if not Path(file_old).exists():
        print(f"Error: File '{file_old}' tidak ditemukan")
        sys.exit(1)
    if not Path(file_new).exists():
        print(f"Error: File '{file_new}' tidak ditemukan")
        sys.exit(1)
    
    print(f"File lama : {file_old}")
    print(f"File baru : {file_new}")
    if key_column:
        print(f"Key column: {key_column}")
    
    # Baca file
    df_old = pd.read_excel(file_old)
    df_new = pd.read_excel(file_new)
    
    print(f"\nFile lama: {len(df_old)} baris, {len(df_old.columns)} kolom")
    print(f"File baru: {len(df_new)} baris, {len(df_new.columns)} kolom")
    
    # Validasi key column
    if key_column:
        if key_column not in df_old.columns:
            print(f"Error: Kolom '{key_column}' tidak ditemukan di file lama")
            print(f"Kolom tersedia: {', '.join(df_old.columns)}")
            sys.exit(1)
        if key_column not in df_new.columns:
            print(f"Error: Kolom '{key_column}' tidak ditemukan di file baru")
            sys.exit(1)
    
    # Bandingkan
    result = compare_dataframes(df_old, df_new, key_column)
    
    # Print report
    print_report(result, key_column)
    
    # Export ke Excel jika diminta
    if output_file:
        export_to_excel(df_old, df_new, result, output_file, key_column)
    
    return result


def main():
    parser = argparse.ArgumentParser(
        description='Membandingkan 2 file Excel dan menampilkan perbedaannya'
    )
    parser.add_argument('file_old', nargs='?', help='File Excel lama')
    parser.add_argument('file_new', nargs='?', help='File Excel baru')
    parser.add_argument('--key', '-k', help='Kolom kunci untuk matching baris')
    parser.add_argument('--output', '-o', help='Export hasil ke file Excel')
    parser.add_argument('--config', '-c', default='config.yaml', help='File konfigurasi YAML')
    
    args = parser.parse_args()
    
    # Load config
    script_dir = Path(__file__).parent
    config = load_config(script_dir / args.config)
    
    # Override dengan CLI (CLI lebih prioritas)
    file_old = args.file_old or config.get('file_old')
    file_new = args.file_new or config.get('file_new')
    key_column = args.key or config.get('key_column')
    output_file = args.output or config.get('output')
    
    # Resolve path
    if file_old and not Path(file_old).is_absolute():
        file_old = str(script_dir / file_old)
    if file_new and not Path(file_new).is_absolute():
        file_new = str(script_dir / file_new)
    if output_file and not Path(output_file).is_absolute():
        output_file = str(script_dir / output_file)
    
    # Validasi
    if not file_old or not file_new:
        print("Error: Kedua file harus diisi (via CLI atau config.yaml)")
        parser.print_help()
        sys.exit(1)
    
    compare_excel(file_old, file_new, key_column, output_file)


if __name__ == "__main__":
    main()
