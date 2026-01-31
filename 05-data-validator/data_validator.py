"""
Data Validator - Validasi data Excel berdasarkan rules di config YAML
Mendukung konfigurasi via YAML dan CLI arguments (hybrid)
"""

import sys
import re
import argparse
import pandas as pd
import yaml
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


FILL_ERROR = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def load_config(config_path: str) -> dict:
    """Load konfigurasi dari file YAML."""
    config_file = Path(config_path)
    if config_file.exists():
        with open(config_file, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f) or {}
    return {}


class Validator:
    """Class untuk validasi data."""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.errors = []  # List of {row, column, value, rule, message}
    
    def validate_required(self, column: str, rule: dict) -> None:
        """Validasi field tidak boleh kosong."""
        for idx, value in self.df[column].items():
            if pd.isna(value) or str(value).strip() == '':
                self.errors.append({
                    'row': idx + 2,  # +2 karena header dan 0-index
                    'column': column,
                    'value': value,
                    'rule': 'required',
                    'message': f'{column} tidak boleh kosong'
                })
    
    def validate_email(self, column: str, rule: dict) -> None:
        """Validasi format email."""
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        for idx, value in self.df[column].items():
            if pd.notna(value) and str(value).strip() != '':
                if not re.match(email_pattern, str(value)):
                    self.errors.append({
                        'row': idx + 2,
                        'column': column,
                        'value': value,
                        'rule': 'email',
                        'message': f'Format email tidak valid: {value}'
                    })
    
    def validate_phone(self, column: str, rule: dict) -> None:
        """Validasi nomor telepon."""
        min_digits = rule.get('min_digits', 10)
        for idx, value in self.df[column].items():
            if pd.notna(value) and str(value).strip() != '':
                digits = re.sub(r'\D', '', str(value))
                if len(digits) < min_digits:
                    self.errors.append({
                        'row': idx + 2,
                        'column': column,
                        'value': value,
                        'rule': 'phone',
                        'message': f'No HP kurang dari {min_digits} digit: {value}'
                    })
    
    def validate_date_range(self, column: str, rule: dict) -> None:
        """Validasi tanggal dalam range."""
        min_date = rule.get('min')
        max_date = rule.get('max')
        
        if min_date:
            min_date = datetime.strptime(min_date, '%Y-%m-%d')
        if max_date:
            max_date = datetime.strptime(max_date, '%Y-%m-%d')
        
        for idx, value in self.df[column].items():
            if pd.notna(value) and str(value).strip() != '':
                try:
                    if isinstance(value, datetime):
                        date_val = value
                    else:
                        date_val = pd.to_datetime(value)
                    
                    if min_date and date_val < min_date:
                        self.errors.append({
                            'row': idx + 2,
                            'column': column,
                            'value': value,
                            'rule': 'date_range',
                            'message': f'Tanggal sebelum {min_date.strftime("%Y-%m-%d")}: {value}'
                        })
                    elif max_date and date_val > max_date:
                        self.errors.append({
                            'row': idx + 2,
                            'column': column,
                            'value': value,
                            'rule': 'date_range',
                            'message': f'Tanggal setelah {max_date.strftime("%Y-%m-%d")}: {value}'
                        })
                except:
                    self.errors.append({
                        'row': idx + 2,
                        'column': column,
                        'value': value,
                        'rule': 'date_range',
                        'message': f'Format tanggal tidak valid: {value}'
                    })
    
    def validate_number_range(self, column: str, rule: dict) -> None:
        """Validasi angka dalam range."""
        min_val = rule.get('min')
        max_val = rule.get('max')
        
        for idx, value in self.df[column].items():
            if pd.notna(value) and str(value).strip() != '':
                try:
                    num_val = float(value)
                    if min_val is not None and num_val < min_val:
                        self.errors.append({
                            'row': idx + 2,
                            'column': column,
                            'value': value,
                            'rule': 'number_range',
                            'message': f'Nilai kurang dari {min_val}: {value}'
                        })
                    elif max_val is not None and num_val > max_val:
                        self.errors.append({
                            'row': idx + 2,
                            'column': column,
                            'value': value,
                            'rule': 'number_range',
                            'message': f'Nilai lebih dari {max_val}: {value}'
                        })
                except:
                    self.errors.append({
                        'row': idx + 2,
                        'column': column,
                        'value': value,
                        'rule': 'number_range',
                        'message': f'Bukan angka valid: {value}'
                    })
    
    def validate_regex(self, column: str, rule: dict) -> None:
        """Validasi dengan regex pattern."""
        pattern = rule.get('pattern')
        message = rule.get('message', f'Tidak sesuai format: {pattern}')
        
        if not pattern:
            return
        
        for idx, value in self.df[column].items():
            if pd.notna(value) and str(value).strip() != '':
                if not re.match(pattern, str(value)):
                    self.errors.append({
                        'row': idx + 2,
                        'column': column,
                        'value': value,
                        'rule': 'regex',
                        'message': message
                    })
    
    def validate_in_list(self, column: str, rule: dict) -> None:
        """Validasi nilai harus dalam list."""
        valid_values = rule.get('values', [])
        
        for idx, value in self.df[column].items():
            if pd.notna(value) and str(value).strip() != '':
                if str(value) not in [str(v) for v in valid_values]:
                    self.errors.append({
                        'row': idx + 2,
                        'column': column,
                        'value': value,
                        'rule': 'in_list',
                        'message': f'Nilai "{value}" tidak valid. Harus salah satu dari: {valid_values}'
                    })
    
    def validate_unique(self, column: str, rule: dict) -> None:
        """Validasi nilai harus unik (tidak duplikat)."""
        seen = {}
        for idx, value in self.df[column].items():
            if pd.notna(value) and str(value).strip() != '':
                str_val = str(value)
                if str_val in seen:
                    self.errors.append({
                        'row': idx + 2,
                        'column': column,
                        'value': value,
                        'rule': 'unique',
                        'message': f'Duplikat dengan baris {seen[str_val]}: {value}'
                    })
                else:
                    seen[str_val] = idx + 2
    
    def validate(self, rules: dict) -> list:
        """Jalankan semua validasi berdasarkan rules."""
        for column, column_rules in rules.items():
            if column not in self.df.columns:
                print(f"  Peringatan: Kolom '{column}' tidak ditemukan, skip validasi")
                continue
            
            for rule in column_rules:
                rule_type = rule.get('type')
                
                if rule_type == 'required':
                    self.validate_required(column, rule)
                elif rule_type == 'email':
                    self.validate_email(column, rule)
                elif rule_type == 'phone':
                    self.validate_phone(column, rule)
                elif rule_type == 'date_range':
                    self.validate_date_range(column, rule)
                elif rule_type == 'number_range':
                    self.validate_number_range(column, rule)
                elif rule_type == 'regex':
                    self.validate_regex(column, rule)
                elif rule_type == 'in_list':
                    self.validate_in_list(column, rule)
                elif rule_type == 'unique':
                    self.validate_unique(column, rule)
        
        return self.errors


def print_report(total_rows: int, errors: list):
    """Print laporan validasi ke console."""
    error_rows = set(e['row'] for e in errors)
    valid_count = total_rows - len(error_rows)
    
    print("\n" + "=" * 50)
    print("LAPORAN VALIDASI")
    print("=" * 50)
    
    print(f"\nRingkasan:")
    print(f"  Total baris : {total_rows}")
    print(f"  Valid       : {valid_count}")
    print(f"  Error       : {len(error_rows)} baris ({len(errors)} masalah)")
    
    if not errors:
        print("\n✓ Semua data VALID!")
        return
    
    print(f"\n--- DETAIL ERROR ({len(errors)}) ---")
    for error in errors[:30]:  # Limit 30
        print(f"  Baris {error['row']}: {error['message']}")
    
    if len(errors) > 30:
        print(f"  ... dan {len(errors) - 30} error lainnya")


def export_to_excel(df: pd.DataFrame, errors: list, output_path: str):
    """Export hasil validasi ke Excel."""
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    error_rows = set(e['row'] for e in errors)
    valid_count = len(df) - len(error_rows)
    
    ws_summary['A1'] = "Laporan Validasi Data"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A3'] = "Total Baris:"
    ws_summary['B3'] = len(df)
    ws_summary['A4'] = "Valid:"
    ws_summary['B4'] = valid_count
    ws_summary['A5'] = "Error:"
    ws_summary['B5'] = len(error_rows)
    ws_summary['A6'] = "Total Masalah:"
    ws_summary['B6'] = len(errors)
    
    # Sheet 2: Detail Error
    if errors:
        ws_errors = wb.create_sheet("Detail Error")
        ws_errors.append(["Baris", "Kolom", "Nilai", "Rule", "Pesan"])
        
        for cell in ws_errors[1]:
            cell.fill = FILL_HEADER
            cell.font = Font(bold=True, color="FFFFFF")
        
        for error in errors:
            ws_errors.append([
                error['row'],
                error['column'],
                str(error['value']) if error['value'] is not None else '',
                error['rule'],
                error['message']
            ])
    
    # Sheet 3: Data dengan highlight error
    ws_data = wb.create_sheet("Data")
    
    # Header
    ws_data.append(list(df.columns))
    for cell in ws_data[1]:
        cell.fill = FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    error_cells = {}  # {(row, col): True}
    for error in errors:
        error_cells[(error['row'], error['column'])] = True
    
    for idx, row in df.iterrows():
        excel_row = idx + 2  # +2 karena header dan 0-index
        row_data = []
        for col in df.columns:
            val = row[col]
            row_data.append(str(val) if pd.notna(val) else '')
        ws_data.append(row_data)
        
        # Highlight error cells
        for col_idx, col in enumerate(df.columns, 1):
            if (excel_row, col) in error_cells:
                ws_data.cell(row=excel_row, column=col_idx).fill = FILL_ERROR
    
    wb.save(output_path)
    print(f"\nLaporan Excel disimpan ke: {output_path}")


def validate_data(input_file: str, rules: dict, output_file: str = None) -> list:
    """Main function untuk validasi data."""
    
    if not Path(input_file).exists():
        print(f"Error: File '{input_file}' tidak ditemukan")
        sys.exit(1)
    
    print(f"Membaca file: {input_file}")
    df = pd.read_excel(input_file)
    print(f"Total baris: {len(df)}")
    print(f"Kolom: {', '.join(df.columns)}")
    
    print("\nMemvalidasi data...")
    validator = Validator(df)
    errors = validator.validate(rules)
    
    print_report(len(df), errors)
    
    if output_file:
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        export_to_excel(df, errors, output_file)
    
    return errors


def main():
    parser = argparse.ArgumentParser(
        description='Validasi data Excel berdasarkan rules di config'
    )
    parser.add_argument('input', nargs='?', help='File Excel input')
    parser.add_argument('--output', '-o', help='Export hasil ke file Excel')
    parser.add_argument('--config', '-c', default='config.yaml', help='File konfigurasi YAML')
    
    args = parser.parse_args()
    
    # Load config
    script_dir = Path(__file__).parent
    config = load_config(script_dir / args.config)
    
    # Override dengan CLI
    input_file = args.input or config.get('input')
    output_file = args.output or config.get('output')
    rules = config.get('rules', {})
    
    # Resolve path
    if input_file and not Path(input_file).is_absolute():
        input_file = str(script_dir / input_file)
    if output_file and not Path(output_file).is_absolute():
        output_file = str(script_dir / output_file)
    
    # Validasi
    if not input_file:
        print("Error: File input harus diisi (via CLI atau config.yaml)")
        parser.print_help()
        sys.exit(1)
    
    if not rules:
        print("Error: Rules validasi harus didefinisikan di config.yaml")
        sys.exit(1)
    
    validate_data(input_file, rules, output_file)


if __name__ == "__main__":
    main()
